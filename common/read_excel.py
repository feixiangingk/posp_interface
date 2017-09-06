#coding:utf-8
import xlrd,sys,copy,random
from win32com.client import Dispatch
import win32com.client
sys.path.append('..')
reload(sys)
sys.setdefaultencoding('utf-8')

from openpyxl import  load_workbook
from datetime import datetime
from interface_init import *
from config.interface_config import InterfaceConfig

class ExecExcel():
    interfaceIndex=None
    interface03_base_case = [1,'QF-POSP-DEPO-CREATE-ACC-0012','10','C11','1003','2','01100300322017071715540002', 'PYZ01',
                             '2017-07-17', '15:54:31', '这大大', '101', '440704196001081942347', '13999705062', 'C-A10->C-B10->ALL->ALL->ALL',
                             '10', '1', 'kduijku@163.com', 'www.baidu.com', '00', '测试', '200', None,
                             '{"tradeNo":"test001","tradeStatus":"30","accountNo":"123456789123456789"}@json', 'y', 'Fail']


    def __init__(self):


        # self.initial=interface_init.initial
        self.file_path="D:\\quarkscript\\posp_interface\\testData\\interface_Data_wh02.xlsx"
        # self.file_path=self.initial.project_path+"\\testData\\interface_Data.xlsx"

        # self.interface_map = self.initial.interfaceConfig.interface_map




        #检查进程是否有excel处于激活状态，有则关闭
        excel=win32com.client.Dispatch('Excel.Application')
        excel.DisplayAlerts=False
        excel.Quit()


    def get_info(self,sheetName):
        #formatting_info=True打开时保留格式
        EL_data=xlrd.open_workbook(self.file_path)
        table=EL_data.sheet_by_name(sheetName)

        #读取接口信息字典
        # interfaceInfo=getattr(self.initial.interfaceConfig,self.interface_map[sheetName])
        # interfaceIndex=self.interface_map[sheetName+"Index"]

        # 不从配置文件读取接口参数列表，而从excel表格直接读取
        interfaceIndex1 = table.row_values(3)
        interfaceIndex1.pop(0)

        #用while循环过滤
        while "" in interfaceIndex1:
            interfaceIndex1.remove("")
        ExecExcel.interfaceIndex=[str(i) for i in interfaceIndex1]


        #确定表格一共多少行数据
        excel_info={}
        for i in xrange(table.nrows):
            #从第四行开始读取excel，前面标题忽略
            if i <=3:
                continue
            rows_info=table.row_values(i)

            excel_info[int(rows_info[0])]={}

            for n in xrange(len(ExecExcel.interfaceIndex)):
                if ExecExcel.interfaceIndex[n] in ['tradeAmount', 'currentSucceedAmount', 'succeedAmount','amount','totalAmount','notifyUrl','email']:
                    excel_info[rows_info[0]][ExecExcel.interfaceIndex[n]] = str(rows_info[n + 1])

                else:
                    excel_info[rows_info[0]][ExecExcel.interfaceIndex[n]]=str(rows_info[n+1]).split('.')[0]



        return excel_info

    #数据驱动读取格式
    def get_info_ddt(self, sheetName):
        # formatting_info=True打开时保留格式
        EL_data = xlrd.open_workbook(self.file_path)
        table = EL_data.sheet_by_name(sheetName)

        # 读取接口信息字典
        # interfaceInfo = getattr(self.initial.interfaceConfig, self.interface_map[sheetName])
        # interfaceIndex = self.interface_map[sheetName + "Index"]

        #不从配置文件读取接口参数列表，而从excel表格直接读取
        interfaceIndex1=table.row_values(3)
        interfaceIndex1.pop(0)

        #用while循环过滤
        # while "" in interfaceIndex1:
        #     interfaceIndex1.remove("")

        #用for循环过滤
        for i in xrange(interfaceIndex1.count("")):
            interfaceIndex1.remove("")

        ExecExcel.interfaceIndex=[str(i) for i in interfaceIndex1]


        # 确定表格一共多少行数据
        excel_info = ['placeholder']

        #用字典映射方法完成 switch case语句
        format_data={"int":lambda x:int(x),
                     "float":lambda x:float(x),
                     "True":lambda x:bool(1),
                     "False":lambda x:bool(0),
                     "list":lambda x:[x],
                     #将字符串转换为json格式
                     "json":lambda x:eval(str(x).replace("\n","").replace("\t",""))}

        for i in xrange(table.nrows):
            # 从第四行开始读取excel，前面标题忽略
            if i <= 3:
                continue
            rows_info = table.row_values(i)

            dict_info={'case_index':int(i-3)}
            for n in xrange(len(ExecExcel.interfaceIndex)):
                if self.interfaceIndex[n] in ['tradeAmount', 'currentSucceedAmount', 'succeedAmount','amount','totalAmount','notifyUrl','email']:
                    dict_info[ExecExcel.interfaceIndex[n]] = str(rows_info[n + 1])
                else:
                    dict_info[ExecExcel.interfaceIndex[n]] = str(rows_info[n + 1]).split('.')[0]

                #第一个if判断值包含@
                if dict_info[ExecExcel.interfaceIndex[n]].find('@')!=-1:
                    #第二个if判断定义的值在我们处理范围内
                    if dict_info[ExecExcel.interfaceIndex[n]].split('@')[1].strip() in ['int','float','list','json','True','Flase']:
                    #用dict.get()方法防止keyerror;用strip()去掉空格
                        try:
                            dict_info[ExecExcel.interfaceIndex[n]]=format_data.get(dict_info[ExecExcel.interfaceIndex[n]].split('@')[1].strip())(dict_info[self.interfaceIndex[n]].split('@')[0])

                        #捕获多个异常的写法
                        except (TypeError,ValueError),e:
                            print "i is {one};n is {two}".format(one=(i-3),two=n+2)
                            # interface_init.initial.logger.info("i is {one};n is {two}".format(one=(i-3),two=n+2))
                            raise e

            excel_info.append(dict_info)

        return excel_info

    def write_info(self,sheetName,result_list):
        EL_data = load_workbook(self.file_path)

        sheet_data = EL_data.get_sheet_by_name(sheetName)

        for case_index,flag,result in result_list:
            row_info=case_index+4
            # interfaceIndex = self.interface_map[sheetName + "Index"]

            #写入执行结果
            sheet_data.cell(coordinate=None,row=row_info,column=len(ExecExcel.interfaceIndex)+2,value=flag)
            #写详情
            sheet_data.cell(coordinate=None,row=row_info,column=len(ExecExcel.interfaceIndex)+3,value=result)

        EL_data.save(self.file_path)

    def create_UseCase(self,sheetName,title_row=4):
        #从配置文件中读取接口信息
        interfaceConfig=InterfaceConfig()
        interface_Name=interfaceConfig.interface_map[sheetName]
        interface_info=getattr(interfaceConfig,interface_Name)
        parmas=interface_info['parmas']
        parmas_dict={}
        for i in  parmas.keys():
            parmas_dict[i]={"type":parmas[i][0],"parmas_len":parmas[i][1],"required_sign":parmas[i][2]}
        # print parmas_dict


        #load需要的sheet数据
        EL_data = load_workbook(self.file_path)
        sheet_data = EL_data.get_sheet_by_name(sheetName)
        # print dir(sheet_data)
        # # print sheet_data.columns()
        # print sheet_data.max_column
        # print sheet_data.max_row
        # print type(sheet_data)
        # print sheet_data['B5'].value
        # print sheet_data.cell(row=5,column=2).value
        # sheet_list=[]
        # for row in sheet_data.iter_rows(min_col=1,max_row=4,max_col=28,min_row=4):
        #     for cell in row:
        #         sheet_list.append(cell.value)
        # print sheet_list

        # sheet_data['B12'].value='sdfk33333'
        # EL_data.save(self.file_path)
        # sheet_data.cell(row=12,column=2,coordinate=None,value='中文试试看')

        # sheet_data.append([1,2,3,4,5,3])
        # sheet_data.append(['sdfsdf',33333,'测试'])
        # EL_data.save(self.file_path)


        '''获取一个字典的映射，key为excel表格表头，value为基础用例的值'''
        title_list = [i.value for i in sheet_data[title_row]]
        base_case = [i.value for i in sheet_data[title_row + 1]]
        data_map = dict(zip(title_list, base_case))
        print parmas_dict

        '''必填元素非空验证'''
        for i in parmas_dict.keys():
            required_sign=parmas_dict[i].get("required_sign")
            if required_sign=="Y":
                parmas_index=title_list.index(i)
                new_case=copy.deepcopy(base_case)
                new_case[parmas_index] = ""
                #用例说明设置
                new_case[-1]="非空验证，参数{one}为空，预期失败".format(one=i)
                #预期结果设置
                tradeStatus_index=title_list.index("re-tradeStatus")
                new_case[tradeStatus_index]=40

                """设置预期返回状态码为601"""
                code_index=title_list.index("re-code")
                new_case[code_index]="601"

                sheet_data.append(new_case)

        '''参数最大长度验证'''
        for i in parmas_dict.keys():
            parmas_len=parmas_dict[i].get("parmas_len")
            parmas_index=title_list.index(i)
            new_case=copy.deepcopy(base_case)
            new_case[parmas_index]="".join([random.choice("abcdefghijklmnopqrstyvw012345678") for k in xrange(parmas_len+1)])
            # 用例说明设置
            new_case[-1] = "参数长度验证，{one}最大长度为{two}，构造{three}传参；预期失败".format(one=i,two=parmas_len,three=new_case[parmas_index])
            # 预期结果设置
            tradeStatus_index = title_list.index("re-tradeStatus")
            new_case[tradeStatus_index] = 40

            """设置预期返回状态码为601"""
            code_index = title_list.index("re-code")
            new_case[code_index] = "601"
            sheet_data.append(new_case)

        '''参数数据类型验证'''
        for i in parmas_dict.keys():
            parmas_type=parmas_dict[i].get("type")
            parmas_index=title_list.index(i)
            new_case=copy.deepcopy(base_case)
            if parmas_type=="string":
                tradeStatus_index = title_list.index("re-tradeStatus")
                new_case[tradeStatus_index] = 40

                """设置预期返回状态码为601"""
                code_index = title_list.index("re-code")
                new_case[code_index] = "601"

                '''新增int类型校验用例'''
                new_case[parmas_index]="1@int"
                new_case[-1]="参数类型校验，{one}应为string型，实际取值int型1，预期失败".format(one=i)
                sheet_data.append(new_case)

                '''新增float类型效验用例'''
                new_case[parmas_index]="2@float"
                new_case[-1]="参数类型校验，{one}应为string型，实际取值float型2.0，预期失败".format(one=i)
                sheet_data.append(new_case)

                '''新增list类型校验用例'''
                new_case[parmas_index]="3@list"
                new_case[-1]="参数类型校验，{one}应为string型，实际取值list型[3]，预期失败".format(one=i)
                sheet_data.append(new_case)
        EL_data.save(self.file_path)



if __name__=="__main__":
    # if isinstance(interface_init.initial,Initialization) !=True:
    #     Init()
    execExcel=ExecExcel()
    list_excel_info=execExcel.get_info_ddt('interface12')
    # print list_excel_info[10]

    execExcel.create_UseCase('interface12')
